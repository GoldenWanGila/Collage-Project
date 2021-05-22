from bs4 import BeautifulSoup
import requests
import openpyxl

wb = openpyxl.Workbook()
row = 1
sheet_position = 0

ws = wb.create_sheet("PTT標題", sheet_position)

for number in range(2400, 2481):
    url = 'https://www.ptt.cc/bbs/Railway/index%s.html' %(number)           # 爬取[情報]

    responses = requests.get(url)
    if responses.status_code == 200:
        soup = BeautifulSoup(responses.text, 'html.parser')

        rents = soup.find_all(class_='r-ent')

        for rent in rents:
            title = rent.find(class_='title').get_text()
            if ("[情報]" in title) and (rent.find(class_='mark').get_text() != 'M') and ('Re' not in title) and ('JR' not in title) and ('新幹線' not in title) and ('日本' not in title):
                title = title.replace('[情報]','')
                date = rent.find(class_='date').get_text().split('/')
                month = '0' + str(int(date[0])) if len(str(int(date[0]))) == 1 else date[0]
                day = '0' + date[1] if len(date[1]) == 1 else date[1]
                year = '2020' if month == '12' else '2021'
                date = year + '/' + month + '/' + day

                ws.cell(row=row, column=1, value=date)
                ws.cell(row=row, column=2, value=title)
                row += 1

    wb.save(r"Data\PTT.xlsx")