import pymysql
import openpyxl
import glob
db=pymysql.connect(host="localhost",user="root",password="Dieark20000115",database="weather")
cursor=db.cursor()

fn = 'Data collector 20210115.xlsx'
wb = openpyxl.load_workbook(fn)
wb.active=0
ws=wb.active

# INSERT指令用到的欄位名稱
column_name="(ObsTime,StnPres,SeaPres,Temperature,Td_Dew_point,RH,WS,WD,WSGust,WDGust,Precp,PrecpHour,SunShine,GloblRad,Visb,UVI,CloudAmount)"

# 欄位名稱
# titles=["ObsTime","StnPres","SeaPres","Temperature","Td_Dew_point","RH","WS","WD","WSGust","WDGust","Precp","PrecpHour","SunShine","GloblRad","Visb","UVI","CloudAmount"]

# 迴圈，每個excel
for filename in glob.glob(r'D:\Dieark\pythons\flask1\Collage-Project\Data\Weather\*.xlsx'):
    # 日期、資料表名稱
    date=filename[-13:-9]+"/"+filename[-9:-7]+"/"+filename[-7:-5]
    # 開啟excel
    fn = 'Data collector '+filename[-13:-5]+'.xlsx'
    wb = openpyxl.load_workbook(fn)
    wb.active=0
    ws=wb.active
    # 判斷資料表是否存在
    sql_find_table="SHOW TABLES LIKE '"+date+"'"
    cursor.execute(sql_find_table)
    data=cursor.fetchall()
    # 資料表不存在，建立資料表
    if len(data)==0:
        sql_new_table="CREATE TABLE `weather`.`"+date+"` ( `ObsTime` VARCHAR(10) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NOT NULL , `StnPres` VARCHAR(10) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NOT NULL , `SeaPres` VARCHAR(10) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NOT NULL , `Temperature` VARCHAR(10) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NOT NULL , `Td_Dew_point` VARCHAR(10) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NOT NULL , `RH` VARCHAR(10) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NOT NULL , `WS` VARCHAR(10) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NOT NULL , `WD` VARCHAR(10) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NOT NULL , `WSGust` VARCHAR(10) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NOT NULL , `WDGust` VARCHAR(10) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NOT NULL , `Precp` VARCHAR(10) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NOT NULL , `PrecpHour` VARCHAR(10) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NOT NULL , `SunShine` VARCHAR(10) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NOT NULL , `GloblRad` VARCHAR(10) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NOT NULL , `Visb` VARCHAR(10) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NOT NULL , `UVI` VARCHAR(10) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NOT NULL , `CloudAmount` VARCHAR(10) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NOT NULL ) ENGINE = MyISAM;"
        cursor.execute(sql_new_table)
    # 插入資料
    for row in ws['A2':'Q25']:
        sql_insert="INSERT INTO `"+date+"` "+column_name+" VALUES ("
        for cell in row:
            sql_insert+=("'"+str(cell.value)+"',")
        sql_insert=sql_insert[:-1]+")"
        sql_insert=sql_insert.replace("\u00a0","")
        print(sql_insert)
        cursor.execute(sql_insert)

db.commit()
print("finish")