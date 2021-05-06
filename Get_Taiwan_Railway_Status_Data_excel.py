from bs4 import BeautifulSoup
import requests
import openpyxl
import time

"""儲存檔案要用的相關參數"""
wb = openpyxl.Workbook()
row = column = 1
sheet_position = 0
direction = ['順行','逆行']

"""欲爬取的站號與站名"""
station_num = ["3230","3280","3300","3350","3340","3360","3390","0900","0960","1000","1020","1080","1100","1210","1250","3160","3470","4080","4220","4340","4400"]
station_name = ["豐原","太原","臺中","潭子","新烏日","彰化","員林","基隆","汐止","臺北","板橋","桃園","中壢","新竹","竹南","苗栗","斗六","嘉義","臺南","新左營","高雄"]

"""用於設定要爬取的日期(rideDate)和儲存檔案要用的日期(fileDate)"""
second = time.time() - 86400
local_time = time.localtime(second)
date = {"year": str(local_time.tm_year), "month": str(local_time.tm_mon), "day": str(local_time.tm_mday)}
rideDate = date["year"] + "/" + date["month"] + "/" + date["day"]
# 為了將檔案格式統一化，所以對month和day進行調整
date["month"] = "0" + date["month"] if len(date["month"]) == 1 else date["month"]
date["day"] = "0" + date["day"] if len(date["day"]) == 1 else date["day"]
fileDate = date["year"] + date["month"] + date["day"]

"""主要執行的程式"""
for index in range(len(station_num)):
    url_byStation = ("https://www.railway.gov.tw/tra-tip-web/tip/tip001/tip112/querybystationblank?rideDate=%s&station=%s-%s" %(rideDate, station_num[index], station_name[index]))
    # print(url_byStation)        # For test

    responses = requests.get(url_byStation)
    soup_byStation = BeautifulSoup(responses.text, "html.parser")

    all_tbody = soup_byStation.find_all(class_="tab-pane")
    # all_tbody[index] : index = 0       --> 順行
    #                    index = 1       --> 逆行

    for direction_index in range(2):
        # direction_index = 0 --> 順行
        # direction_index = 1 --> 逆行
        ws = wb.create_sheet("%s(%s)" % (station_name[index], direction[direction_index]), sheet_position)
        sheet_position += 1
        all_item = all_tbody[direction_index].find_all("td")
        # all_item[index] : index = 6k      item --> 排序編號
        #                   index = 1+6k    item --> 車種+車次(起點->終點)
        #                   index = 2+6k    item --> 預計進站時間
        #                   index = 3+6k    item --> 終點站站名 
        #                   index = 4+6k    item --> 此班列車行駛狀態
        #                   index = 5+6k    item --> 台鐵實際進站狀態
        for item in all_item:
            if column == 7:
                status = ''.join([x for x in item.get_text() if x.isdigit()])
                if status == '':
                    status = '0'
                ws.cell(row = row, column = column, value = status)
                row += 1
                column = 0
            else:
                if column != 2:
                    ws.cell(row = row, column = column, value = item.get_text())
                else:
                    """設定車種儲存"""
                    train_catg = ''
                    train_catg_temp = []
                    for x in item.get_text():
                            if x != '\n':
                                if x.isdigit() == False:
                                    train_catg_temp += [x]
                                else:
                                    train_catg = ''.join(train_catg_temp)
                                    break
                    """設定車次儲存"""
                    train_num = ''.join([x for x in item.get_text() if x.isdigit()])
                    
                    ws.cell(row = row, column = 2, value = train_catg)
                    ws.cell(row = row, column = 3, value = train_num)
                    column = 3
            column += 1
        row = 1

wb.save(r"TestData\Railway\Data collector %s.xlsx" %fileDate)
