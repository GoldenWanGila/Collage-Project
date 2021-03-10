from bs4 import BeautifulSoup
import requests
import openpyxl
import time

column_indexs = ['ObsTime','StnPres','SeaPres','Temperature','Td_dew_point','RH','WS','WD','WSGust','WDGust','Precp',
                'PrecpHour','SunShine','GloblRad','Visb','UVI','CloudAmount']
station_name = ['台中','豐原','潭子','烏日','花壇','員林']
station_num = ['467490','C0F9M0','C0F9O0','C0F9S0','C0G910','C0G650']

"""儲存檔案要用的相關參數"""
wb = openpyxl.Workbook()
row = 1
column = 0

"""用於設定要爬取的日期(crawlDate)和儲存檔案要用的日期(fileDate)"""
second = time.time() - 86400
local_time = time.localtime(second)
date = {"year": str(local_time.tm_year), "month": str(local_time.tm_mon), "day": str(local_time.tm_mday)}
date["month"] = "0" + date["month"] if len(date["month"]) == 1 else date["month"]
date["day"] = "0" + date["day"] if len(date["day"]) == 1 else date["day"]
fileDate = date["year"] + date["month"] + date["day"]
crawlDate = date["year"] + "-" + date["month"] + "-" + date["day"]

if (local_time.tm_hour>=12):
    for index1 in range(len(station_num)):
        url = 'https://e-service.cwb.gov.tw/HistoryDataQuery/DayDataController.do?command=viewMain&station=%s&stname=%s&datepicker=%s' %(station_num[index1], '%25E8%2587%25BA%25E4%25B8%25AD', crawlDate)

        sheet_position = index1
        ws = wb.create_sheet(station_name[index1], sheet_position)
        for item in column_indexs:
            column += 1
            ws.cell(row=row, column=column, value=item)
        row = 2
        column = 0

        response = requests.get(url)
        soup = BeautifulSoup(response.text, 'html.parser')

        all_tr = soup.find('tbody').find_all('tr')

        for index2 in range(3, 27):
            td_in_tr = all_tr[index2].find_all('td')
            for td in td_in_tr:
                column += 1
                if column == 1:
                    ws.cell(row=row, column=column, value=td.get_text()+':00')
                else:
                    ws.cell(row=row, column=column, value=td.get_text())
            if column == 17:
                row += 1
                column = 0
            if row == 26:
                row = 1

    wb.save(r"Data\Weather\Data collector %s.xlsx" %fileDate)
else:
    print("\n現在還不是爬取的時候，等到12:00過後再來\n")