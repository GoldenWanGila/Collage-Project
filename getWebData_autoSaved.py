from bs4 import BeautifulSoup
import requests
import openpyxl
import time

wb = openpyxl.Workbook()
local_time = time.localtime()

date = {"year": "", "month":"", "day":""}
date["year"] = str(local_time.tm_year)                          # 更改這邊就可以改日期
date["month"] = str(local_time.tm_mon)                          # 更改這邊就可以改日期
date["day"] = str(local_time.tm_mday - 1)                       # 更改這邊就可以改日期
rideDate = date["year"] + "/" + date["month"] + "/" + date["day"]
fileDate = date["year"] + date["month"] + date["day"]

station_num = ["3230","3280","3300","3350","3340","3360","3390"]
station_name = ["豐原","太原","臺中","潭子","新烏日","彰化","員林"]

row = 1
column = 0
sheet_position = 0

for index in range(7):
    url = ("https://www.railway.gov.tw/tra-tip-web/tip/tip001/tip112/querybystationblank?rideDate=%s&station=%s-%s" %(rideDate, station_num[index], station_name[index]))
    # print(url)        # For test

    responses = requests.get(url)
    soup = BeautifulSoup(responses.text, "html.parser")

    all_tbody = soup.find_all(class_="tab-pane")

    # 順行
    ws = wb.create_sheet("%s(順行)" % station_name[index], sheet_position)
    all_item = all_tbody[0].find_all("td")
    # all_item[index] : index = 6C      --> 排序編號
    #                   index = 1+6C    --> 車種+車次(起點->終點)
    #                   index = 2+6C    --> 預計進站時間
    #                   index = 3+6C    --> 終點站站名 
    #                   index = 4+6C    --> 此班列車行駛狀態
    #                   index = 5+6C    --> 台鐵實際進站狀態
    for item in all_item:
        column += 1
        ws.cell(row = row, column = column, value = item.get_text())
        if column == 6:
            row += 1
            column = 0
    row = 1

    #逆行
    ws = wb.create_sheet("%s(逆行)" % station_name[index], sheet_position+1)
    all_item = all_tbody[1].find_all("td")
    sheet_position += 2
    
    for item in all_item:
        column += 1
        ws.cell(row = row, column = column, value = item.get_text())
        if column == 6:
            row += 1
            column = 0
    row = 1

wb.save(r"GoldenWanGila\Collage-Project\Data\Data collector %s.xlsx" %fileDate)
