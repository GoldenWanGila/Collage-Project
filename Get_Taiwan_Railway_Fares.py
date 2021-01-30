from Authorization import Auth
import requests
from bs4 import BeautifulSoup
import openpyxl

"""儲存檔案要用的相關參數"""
wb = openpyxl.Workbook()
ws = wb.create_sheet("一般成人票(出發站皆為臺中)",0)
ws.cell(row=1,column=1,value="迄點車站")
ws.cell(row=1,column=2,value="迄點車站代碼")
ws.cell(row=1,column=3,value="乘車距離(km)")
ws.cell(row=1,column=4,value="區間票價(元台幣)")
ws.cell(row=1,column=5,value="莒光票價(元台幣)")
ws.cell(row=1,column=6,value="普悠瑪/太魯閣/自強票價(元台幣)")

row = 2

fareRates = {"區間":1.46,"莒光":1.75,"自強":2.27}

APPID = "ea4357a24c3d4072a3b2d1b538f19984"
APPKEY = '5krzdQvXot_vms6ciqpR4emxoHE'

station_dict = {"3300":"臺中","3230":"豐原","3280":"太原","3350":"潭子","3340":"新烏日","3360":"彰化","3390":"員林","0900":"基隆","0960":"汐止","1000":"臺北","1020":"板橋","1080":"桃園","1100":"中壢","1210":"新竹","1250":"竹南","3160":"苗栗","3470":"斗六","4080":"嘉義","4220":"臺南","4340":"新左營","4400":"高雄"}
station_num = ["3230","3280","3350","3340","3360","3390","0900","0960","1000","1020","1080","1100","1210","1250","3160","3470","4080","4220","4340","4400"]



auth = Auth(APPID, APPKEY)
myheader = auth.get_header()

for index in range(len(station_num)):
    OriginStation = "3300"
    DestinationStation = EndStation = station_num[index]

    if OriginStation < DestinationStation:
        temp = OriginStation
        OriginStation = DestinationStation
        DestinationStation = temp

    url = ('https://ptx.transportdata.tw/MOTC/v3/Rail/TRA/ODFare/%s/to/%s?$select=TravelDistance&$filter=((Direction)eq(0))and((TrainType)eq(1))&$format=XML'%(OriginStation, DestinationStation))
    responses = requests.get(url, headers = myheader)
    soup = BeautifulSoup(responses.text, 'html.parser')

    dist = float(soup.find('traveldistance').get_text())

    result = [station_dict[str(EndStation)], EndStation, dist]
    for fare_rate in fareRates:
        result.append(int(dist * fareRates[fare_rate]))

    for column in range(1,7):
        ws.cell(row=row,column=column,value=result[column-1])
    
    row += 1

wb.save(r"Data\TaiwanRailwayFares.xlsx")